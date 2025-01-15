import openpyxl
import csv
from typing import Dict, Tuple
import os


def excel_to_csv(excel_file_path: str, csv_file_path: str, max_width: int = 400, max_height: int = 100):
    # Check if the Excel file exists
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active

    # Define mappings
    terrain_map = {
        'B': 'BARE',
        'S': 'SPARSE_VEG',
        'D': 'DENSE_VEG',
        'W': 'WOODS',
        'T': 'STRUCTURE'
    }
    elevation_map = {
        'G': 'GROUND_LEVEL',
        'E': 'ELEVATED_LEVEL',
        'L': 'LOWER_LEVEL'
    }

    # Check if the Excel sheet dimensions match the environment size
    if sheet.max_column > max_width or sheet.max_row > max_height:
        print(f"Warning: Excel sheet dimensions ({sheet.max_column}x{sheet.max_row}) "
              f"exceed the environment size ({max_width}x{max_height}).")
        print("Extra cells will be ignored.")
    elif sheet.max_column < max_width or sheet.max_row < max_height:
        print(f"Warning: Excel sheet dimensions ({sheet.max_column}x{sheet.max_row}) "
              f"are smaller than the environment size ({max_width}x{max_height}).")
        print("Missing cells will be filled with default values (BARE, GROUND_LEVEL).")

    # Ensure the directory for the CSV file exists
    os.makedirs(os.path.dirname(csv_file_path), exist_ok=True)

    # Open CSV file for writing
    with open(csv_file_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['x', 'y', 'terrain_type', 'elevation_type'])  # Write header

        # Iterate through cells
        for row in range(1, max_height + 1):
            for col in range(1, max_width + 1):
                if row <= sheet.max_row and col <= sheet.max_column:
                    cell_value = sheet.cell(row=row, column=col).value
                else:
                    cell_value = None

                if cell_value:
                    terrain = terrain_map.get(cell_value[0], 'BARE')
                    elevation = elevation_map.get(cell_value[1], 'GROUND_LEVEL')
                else:
                    terrain = 'BARE'
                    elevation = 'GROUND_LEVEL'

                writer.writerow([col - 1, row - 1, terrain, elevation])

    print(f"CSV file '{csv_file_path}' has been created with dimensions {max_width}x{max_height}.")


if __name__ == "__main__":
    # Get the current script's directory
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Construct full paths for input and output files
    excel_file = os.path.join(script_dir, "map_design.xlsx")
    csv_file = os.path.join(script_dir, "generated_map.csv")

    # Check if the Excel file exists
    if not os.path.exists(excel_file):
        print(f"Excel file not found: {excel_file}")
        print("Please ensure the Excel file is in the same directory as this script.")
    else:
        print(f"Converting {excel_file} to CSV...")
        excel_to_csv(excel_file, csv_file)
